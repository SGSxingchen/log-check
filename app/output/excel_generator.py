"""Excel报表生成模块"""
import re
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def generate_excel(daily_stats, dialogue_contents, file_path="对话统计.xlsx"):
    """生成包含统计数据和每个角色对话内容的Excel文件"""
    wb = openpyxl.Workbook()
    
    # 删除默认创建的Sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # 设置单元格样式
    header_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    day_header_fill = PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid")
    centered = Alignment(horizontal='center')
    wrapped_text = Alignment(wrap_text=True, vertical='top')
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    ooc_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # 创建总体统计工作表
    ws_summary = wb.create_sheet(title="统计")
    
    # 计算总计数据并检查是否有参与时长
    total_stats = defaultdict(lambda: {
        'messages': 0,
        'chars': 0,
        'ooc_messages': 0,
        'ooc_chars': 0,
        'participation_minutes': 0
    })
    has_participation_time = False
    for day_stats in daily_stats.values():
        for speaker, stats in day_stats.items():
            total_stats[speaker]['messages'] += stats['messages']
            total_stats[speaker]['chars'] += stats['chars']
            total_stats[speaker]['ooc_messages'] += stats['ooc_messages']
            total_stats[speaker]['ooc_chars'] += stats['ooc_chars']
            total_stats[speaker]['participation_minutes'] += stats.get('participation_minutes', 0)
            if stats.get('participation_minutes', 0) > 0:
                has_participation_time = True

    # 根据是否有参与时长设置列宽和表头
    num_cols = 7 if has_participation_time else 6
    ws_summary.column_dimensions['A'].width = 30
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws_summary.column_dimensions[col_letter].width = 12

    # 添加标题
    ws_summary['A1'] = "角色对话总体统计"
    ws_summary['A1'].font = title_font
    ws_summary.merge_cells(f'A1:{chr(64+num_cols)}1')
    ws_summary['A1'].alignment = Alignment(horizontal='center')

    # 添加总体统计表头
    total_headers = ["角色名", "总发言数", "总字数", "平均字数", "场外次数", "场外字数"]
    if has_participation_time:
        total_headers.append("参与时长")

    for col, header in enumerate(total_headers, 1):
        cell = ws_summary.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centered
        cell.border = border

    # 添加总体统计数据
    row = 4
    for speaker, stats in sorted(total_stats.items(),
                              key=lambda x: x[1]['messages'],
                              reverse=True):
        avg_chars = stats['chars'] / stats['messages'] if stats['messages'] > 0 else 0

        ws_summary.cell(row=row, column=1).value = speaker
        ws_summary.cell(row=row, column=2).value = stats['messages']
        ws_summary.cell(row=row, column=3).value = stats['chars']
        ws_summary.cell(row=row, column=4).value = f"{avg_chars:.1f}"
        ws_summary.cell(row=row, column=5).value = stats['ooc_messages']
        ws_summary.cell(row=row, column=6).value = stats['ooc_chars']

        col_count = 6
        if has_participation_time:
            hours = stats['participation_minutes'] // 60
            minutes = stats['participation_minutes'] % 60
            time_str = f"{hours}小时{minutes}分钟" if hours > 0 else f"{minutes}分钟"
            ws_summary.cell(row=row, column=7).value = time_str
            col_count = 7

        # 添加边框和居中
        for col in range(1, col_count + 1):
            cell = ws_summary.cell(row=row, column=col)
            cell.border = border
            if col > 1:  # 第一列左对齐，其他居中
                cell.alignment = centered

        row += 1

    # 添加空行
    row += 2
    
    # 在同一工作表中添加每日统计数据
    for day, day_stats in sorted(daily_stats.items()):
        # 添加每日标题
        day_title_cell = ws_summary.cell(row=row, column=1)
        day_title_cell.value = f"第{day}天对话统计"
        day_title_cell.font = title_font
        ws_summary.merge_cells(f'A{row}:{chr(64+num_cols)}{row}')
        day_title_cell.alignment = Alignment(horizontal='center')
        row += 1

        # 添加表头
        day_headers = ["角色名", "发言数", "字数", "平均字数", "场外次数", "场外字数"]
        if has_participation_time:
            day_headers.append("参与时长")

        for col, header in enumerate(day_headers, 1):
            cell = ws_summary.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = day_header_fill
            cell.alignment = centered
            cell.border = border
        row += 1

        # 添加数据
        daily_total = defaultdict(int)

        for speaker, stats in sorted(day_stats.items(),
                                  key=lambda x: x[1]['messages'],
                                  reverse=True):
            avg_chars = stats['chars'] / stats['messages'] if stats['messages'] > 0 else 0

            ws_summary.cell(row=row, column=1).value = speaker
            ws_summary.cell(row=row, column=2).value = stats['messages']
            ws_summary.cell(row=row, column=3).value = stats['chars']
            ws_summary.cell(row=row, column=4).value = f"{avg_chars:.1f}"
            ws_summary.cell(row=row, column=5).value = stats['ooc_messages']
            ws_summary.cell(row=row, column=6).value = stats['ooc_chars']

            col_count = 6
            if has_participation_time:
                hours = stats.get('participation_minutes', 0) // 60
                minutes = stats.get('participation_minutes', 0) % 60
                time_str = f"{hours}小时{minutes}分钟" if hours > 0 else f"{minutes}分钟"
                ws_summary.cell(row=row, column=7).value = time_str
                col_count = 7

            # 添加边框和居中
            for col in range(1, col_count + 1):
                cell = ws_summary.cell(row=row, column=col)
                cell.border = border
                if col > 1:  # 第一列左对齐，其他居中
                    cell.alignment = centered
            
            # 累计当日总数
            daily_total['messages'] += stats['messages']
            daily_total['chars'] += stats['chars']
            daily_total['ooc_messages'] += stats['ooc_messages']
            daily_total['ooc_chars'] += stats['ooc_chars']
            daily_total['participation_minutes'] += stats.get('participation_minutes', 0)

            row += 1

        # 添加当日总计行
        avg_chars_total = daily_total['chars'] / daily_total['messages'] if daily_total['messages'] > 0 else 0

        ws_summary.cell(row=row, column=1).value = "当日总计"
        ws_summary.cell(row=row, column=1).font = header_font
        ws_summary.cell(row=row, column=2).value = daily_total['messages']
        ws_summary.cell(row=row, column=3).value = daily_total['chars']
        ws_summary.cell(row=row, column=4).value = f"{avg_chars_total:.1f}"
        ws_summary.cell(row=row, column=5).value = daily_total['ooc_messages']
        ws_summary.cell(row=row, column=6).value = daily_total['ooc_chars']

        col_count_total = 6
        if has_participation_time:
            hours = daily_total['participation_minutes'] // 60
            minutes = daily_total['participation_minutes'] % 60
            time_str = f"{hours}小时{minutes}分钟" if hours > 0 else f"{minutes}分钟"
            ws_summary.cell(row=row, column=7).value = time_str
            col_count_total = 7

        # 添加边框和背景色
        for col in range(1, col_count_total + 1):
            cell = ws_summary.cell(row=row, column=col)
            cell.border = border
            cell.fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
            if col > 1:  # 第一列左对齐，其他居中
                cell.alignment = centered
        
        # 添加空行
        row += 2
    
    # 为每个角色创建对话内容工作表
    for speaker, dialogues in sorted(dialogue_contents.items()):
        print(f"Processing speaker: {speaker}")
        if not dialogues:  # 跳过没有对话的角色
            continue
            
        # 创建工作表（名称不能超过31个字符，且不能包含特殊字符）
        safe_name = re.sub(r'[\[\]\:\*\?\/\\]', '', speaker)[:31]
        ws = wb.create_sheet(title=safe_name)
        
        # 设置列宽
        ws.column_dimensions['A'].width = 40  # 日期/时间戳
        ws.column_dimensions['B'].width = 70  # 对话内容
        
        # 添加标题
        ws['A1'] = f"{speaker} 对话内容"
        ws['A1'].font = title_font
        ws.merge_cells('A1:B1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # 添加表头
        headers = ["时间", "对话内容"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = centered
            cell.border = border


        # 添加对话数据（按天和时间排序）
        row = 4
        for dialogue in sorted(dialogues, key=lambda x: (x['day'], x.get('timestamp', ''))):
            # 时间戳
            timestamp_cell = ws.cell(row=row, column=1)
            timestamp_cell.value = dialogue.get('timestamp', f"第{dialogue['day']}天")
            timestamp_cell.border = border
            timestamp_cell.alignment = centered
            
            # 对话内容
            content_cell = ws.cell(row=row, column=2)
            content_cell.value = dialogue['content']
            content_cell.border = border
            content_cell.alignment = wrapped_text
            
            # 如果是场外内容，使用不同背景色
            if dialogue['is_ooc']:
                timestamp_cell.fill = ooc_fill
                content_cell.fill = ooc_fill
            
            row += 1
    
    # 保存Excel文件
    try:
        wb.save(file_path)
        print(f"Excel文件已生成：{file_path}")
    except Exception as e:
        print(f"生成Excel文件时出错：{e}")
