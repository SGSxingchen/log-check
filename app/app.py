import re
import csv
import os
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def count_char_weight(text):
    """计算文本的加权字数，汉字计4，其他字符计1"""
    weight = 0
    for char in text:
        if '\u4e00' <= char <= '\u9fff':  # 判断是否为汉字
            weight += 4
        else:
            weight += 1
    return weight

def detect_format(file_path):
    """检测日志文件的格式类型"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue

                # 格式1: 用户名(ID) 年-月-日 时间
                if re.match(r'.+?\(\d+\) \d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}', line):
                    return "format1"

                # 格式2: 用户名(ID) 年/月/日 时间
                if re.match(r'.+?\(\d+\) \d{4}/\d{2}/\d{2} \d{2}:\d{2}:\d{2}', line):
                    return "format2"

                # 格式3: 时间 <用户名>消息内容 (无冒号)
                if re.match(r'\d{2}:\d{2}:\d{2} <([^>]+)>(.*)', line):
                    return "format3"

                # 格式4: 时间<用户名>:消息内容
                if re.match(r'(?:\d{2}:\d{2}:\d{2})?<([^>]+)>:(.*)', line):
                    return "format4"
    except FileNotFoundError:
        print(f"警告: 文件 {file_path} 未找到")

    # 默认返回格式1
    return "format1"

def count_dialogue_from_file(file_path="log.txt"):
    """从文件中统计对话信息，支持四种格式，并返回每个角色的具体对话内容"""
    daily_stats = defaultdict(lambda: defaultdict(lambda: {
        'messages': 0, 
        'chars': 0,
        'ooc_messages': 0,
        'ooc_chars': 0
    }))
    
    # 存储每个角色的具体对话内容
    dialogue_contents = defaultdict(list)
    
    current_day = 1
    current_speaker = None
    current_message = []
    current_date = None
    current_timestamp = None
    
    # 检测文件格式
    format_type = detect_format(file_path)
    print(f"检测到{format_type}格式日志文件")
    
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue

                if format_type in ["format1", "format2"]:  # 处理格式1和格式2（带日期时间戳）
                    # 根据格式类型获取日期匹配模式
                    date_pattern = r'\d{4}-\d{2}-\d{2}' if format_type == "format1" else r'\d{4}/\d{2}/\d{2}'
                    
                    # 检查是否为新日期（用于分天）
                    date_match = re.search(date_pattern, line)
                    if date_match and current_date != date_match.group(0):
                        if current_date is not None:
                            # 处理切换日期前的最后一条消息
                            if current_speaker and current_message:
                                content = '\n'.join(current_message)
                                is_ooc = bool(re.search(r'[（(]', content))
                                if is_ooc:
                                    daily_stats[current_day][current_speaker]['ooc_messages'] += 1
                                    daily_stats[current_day][current_speaker]['ooc_chars'] += count_char_weight(content)
                                else:
                                    daily_stats[current_day][current_speaker]['messages'] += 1
                                    daily_stats[current_day][current_speaker]['chars'] += count_char_weight(content)
                                    
                                # 保存对话内容
                                dialogue_contents[current_speaker].append({
                                    'day': current_day,
                                    'date': current_date,
                                    'content': content,
                                    'is_ooc': is_ooc,
                                    'timestamp': current_timestamp
                                })
                            # 切换到新的一天
                            current_day += 1
                        current_date = date_match.group(0)
                    
                    # 匹配格式1或格式2的发言行，根据格式类型使用不同的正则
                    speaker_pattern = r'(.+?)\((\d+)\) \d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}' if format_type == "format1" else r'(.+?)\((\d+)\) \d{4}/\d{2}/\d{2} \d{2}:\d{2}:\d{2}'
                    speaker_match = re.match(speaker_pattern, line)
                    if speaker_match:
                        # print(speaker_match)
                        # 处理前一条消息
                        if current_speaker and current_message:
                            content = '\n'.join(current_message)
                            is_ooc = bool(re.search(r'[（(]', content))
                            if is_ooc:
                                daily_stats[current_day][current_speaker]['ooc_messages'] += 1
                                daily_stats[current_day][current_speaker]['ooc_chars'] += count_char_weight(content)
                            else:
                                print(f"{current_speaker}:{daily_stats[current_day][current_speaker]}")
                                daily_stats[current_day][current_speaker]['messages'] += 1
                                daily_stats[current_day][current_speaker]['chars'] += count_char_weight(content)
                            # print(current_date)
                            # 保存对话内容
                            dialogue_contents[current_speaker].append({
                                'day': current_day,
                                'date': current_date,
                                'content': content,
                                'is_ooc': is_ooc,
                                'timestamp': current_timestamp
                            })
                            # print(f"{current_speaker}:{dialogue_contents[current_speaker]}")
                        
                        current_speaker = speaker_match.group(1)
                        current_timestamp = line
                        current_message = []
                    else:
                        # 如果已有当前发言者，则这行是消息内容
                        if current_speaker:
                            current_message.append(line)
                
                else:  # 处理格式3和格式4（基于时间戳的格式）
                    # 检查是否是日志分隔符，表示新的一天
                    if ('.log off' in line or '.logoff' in line or 
                        '。log off' in line or '。logoff' in line or 
                        '/log off' in line or '/logoff' in line or 
                        '===日志开始===' in line or '===日志结束===' in line or 
                        re.match(r'={3,}.*={3,}', line)):
                        # 处理当前未处理的消息
                        if current_speaker and current_message:
                            content = '\n'.join(current_message)
                            is_ooc = bool(re.search(r'[（(]', content))
                            if is_ooc:
                                daily_stats[current_day][current_speaker]['ooc_messages'] += 1
                                daily_stats[current_day][current_speaker]['ooc_chars'] += count_char_weight(content)
                            else:
                                daily_stats[current_day][current_speaker]['messages'] += 1
                                daily_stats[current_day][current_speaker]['chars'] += count_char_weight(content)
                            
                            # 保存对话内容
                            dialogue_contents[current_speaker].append({
                                'day': current_day,
                                'date': None,  # 格式3和格式4没有日期字段
                                'content': content,
                                'is_ooc': is_ooc,
                                'timestamp': f"第{current_day}天"  # 使用天数作为时间戳
                            })
                        current_speaker = None
                        current_message = []
                        current_day += 1
                        continue
                    
                    if format_type == "format3":
                        # 处理格式3：时间 <用户名>消息内容
                        match = re.match(r'(\d{2}:\d{2}:\d{2}) <([^>]+)>(.*)', line)
                        if match:
                            # 检查时间跨天情况
                            current_time = match.group(1)
                            current_hour = int(current_time[:2])
                            
                            # 如果有前一个时间戳，检查是否跨天
                            if 'prev_timestamp' in locals() and prev_timestamp:
                                prev_hour = int(prev_timestamp[:2])
                                # 如果前一个时间是23:xx，当前时间是00:xx-05:xx，认为是跨天
                                if prev_hour >= 23 and current_hour <= 5:
                                    # 处理前一条消息
                                    if current_speaker and current_message:
                                        content = '\n'.join(current_message)
                                        is_ooc = bool(re.search(r'[（(]', content))
                                        if is_ooc:
                                            daily_stats[current_day][current_speaker]['ooc_messages'] += 1
                                            daily_stats[current_day][current_speaker]['ooc_chars'] += count_char_weight(content)
                                        else:
                                            daily_stats[current_day][current_speaker]['messages'] += 1
                                            daily_stats[current_day][current_speaker]['chars'] += count_char_weight(content)
                                        
                                        # 保存对话内容
                                        dialogue_contents[current_speaker].append({
                                            'day': current_day,
                                            'date': None,  # 格式3和格式4没有日期字段
                                            'content': content,
                                            'is_ooc': is_ooc,
                                            'timestamp': f"第{current_day}天 {prev_timestamp}"
                                        })
                                    # 跨天，增加天数
                                    current_day += 1
                                    current_speaker = None
                                    current_message = []
                            
                            # 处理前一条消息（如果不是跨天情况）
                            if current_speaker and current_message:
                                content = '\n'.join(current_message)
                                is_ooc = bool(re.search(r'[（(]', content))
                                if is_ooc:
                                    daily_stats[current_day][current_speaker]['ooc_messages'] += 1
                                    daily_stats[current_day][current_speaker]['ooc_chars'] += count_char_weight(content)
                                else:
                                    daily_stats[current_day][current_speaker]['messages'] += 1
                                    daily_stats[current_day][current_speaker]['chars'] += count_char_weight(content)
                                
                                # 保存对话内容
                                dialogue_contents[current_speaker].append({
                                    'day': current_day,
                                    'date': None,  # 格式3和格式4没有日期字段
                                    'content': content,
                                    'is_ooc': is_ooc,
                                    'timestamp': f"第{current_day}天 {prev_timestamp if 'prev_timestamp' in locals() else ''}"
                                })
                            
                            prev_timestamp = match.group(1)
                            current_speaker = match.group(2)
                            current_message = [match.group(3).strip()]
                        else:
                            if current_speaker and line:
                                current_message.append(line)
                    else:
                        # 匹配格式4发言行：时间<用户名>:消息内容
                        match = re.match(r'(?:(\d{2}:\d{2}:\d{2}))?<([^>]+)>:(.*)', line)
                        if match:
                            current_time = match.group(1) if match.group(1) else ""
                            
                            # 如果有时间戳，检查是否跨天
                            if current_time and 'prev_timestamp' in locals() and prev_timestamp:
                                current_hour = int(current_time[:2])
                                prev_hour = int(prev_timestamp[:2])
                                # 如果前一个时间是23:xx，当前时间是00:xx-05:xx，认为是跨天
                                if prev_hour >= 23 and current_hour <= 5:
                                    # 处理前一条消息
                                    if current_speaker and current_message:
                                        content = '\n'.join(current_message)
                                        is_ooc = bool(re.search(r'[（(]', content))
                                        if is_ooc:
                                            daily_stats[current_day][current_speaker]['ooc_messages'] += 1
                                            daily_stats[current_day][current_speaker]['ooc_chars'] += count_char_weight(content)
                                        else:
                                            daily_stats[current_day][current_speaker]['messages'] += 1
                                            daily_stats[current_day][current_speaker]['chars'] += count_char_weight(content)
                                        
                                        # 保存对话内容
                                        dialogue_contents[current_speaker].append({
                                            'day': current_day,
                                            'date': None,  # 格式3和格式4没有日期字段
                                            'content': content,
                                            'is_ooc': is_ooc,
                                            'timestamp': f"第{current_day}天 {prev_timestamp}"
                                        })
                                    # 跨天，增加天数
                                    current_day += 1
                                    current_speaker = None
                                    current_message = []
                            
                            # 处理前一条消息（如果不是跨天情况）
                            if current_speaker and current_message:
                                content = '\n'.join(current_message)
                                is_ooc = bool(re.search(r'[（(]', content))
                                if is_ooc:
                                    daily_stats[current_day][current_speaker]['ooc_messages'] += 1
                                    daily_stats[current_day][current_speaker]['ooc_chars'] += count_char_weight(content)
                                else:
                                    daily_stats[current_day][current_speaker]['messages'] += 1
                                    daily_stats[current_day][current_speaker]['chars'] += count_char_weight(content)
                                
                                # 保存对话内容
                                dialogue_contents[current_speaker].append({
                                    'day': current_day,
                                    'date': None,  # 格式3和格式4没有日期字段
                                    'content': content,
                                    'is_ooc': is_ooc,
                                    'timestamp': f"第{current_day}天 {prev_timestamp if 'prev_timestamp' in locals() else ''}"
                                })
                            
                            prev_timestamp = match.group(1) if match.group(1) else ""
                            current_speaker = match.group(2)
                            current_message = [match.group(3).strip()]
                        else:
                            if current_speaker and line:
                                current_message.append(line)

    except FileNotFoundError:
        print(f"警告: 文件 {file_path} 未找到")
        
    # 处理最后一条消息
    if current_speaker and current_message:
        content = '\n'.join(current_message)
        is_ooc = bool(re.search(r'[（(]', content))
        
        if is_ooc:
            daily_stats[current_day][current_speaker]['ooc_messages'] += 1
            daily_stats[current_day][current_speaker]['ooc_chars'] += count_char_weight(content)
        else:
            daily_stats[current_day][current_speaker]['messages'] += 1
            daily_stats[current_day][current_speaker]['chars'] += count_char_weight(content)
        
        # 保存最后一条消息
        dialogue_contents[current_speaker].append({
            'day': current_day,
            'date': current_date,
            'content': content,
            'is_ooc': is_ooc,
            'timestamp': f"第{current_day}天" if format_type in ["format3", "format4"] else f"第{current_day}天 {current_date}"
        })

    return daily_stats, dialogue_contents

def generate_excel(daily_stats, dialogue_contents, file_path="对话统计.xlsx"):
    """生成包含统计数据和每个角色对话内容的Excel文件"""
    wb = openpyxl.Workbook()
    
    # 删除默认创建的Sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # 设置单元格样式
    header_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    day_header_fill = PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid")
    centered = Alignment(horizontal='center')
    wrapped_text = Alignment(wrap_text=True, vertical='top')
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    ooc_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # 创建总体统计工作表
    ws_summary = wb.create_sheet(title="统计")
    
    # 设置总体统计表的列宽
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 12
    ws_summary.column_dimensions['C'].width = 12
    ws_summary.column_dimensions['D'].width = 12
    ws_summary.column_dimensions['E'].width = 12
    ws_summary.column_dimensions['F'].width = 12
    
    # 添加标题
    ws_summary['A1'] = "角色对话总体统计"
    ws_summary['A1'].font = title_font
    ws_summary.merge_cells('A1:F1')
    ws_summary['A1'].alignment = Alignment(horizontal='center')
    
    # 添加总体统计表头
    total_headers = ["角色名", "总发言数", "总字数", "平均字数", "场外次数", "场外字数"]
    for col, header in enumerate(total_headers, 1):
        cell = ws_summary.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centered
        cell.border = border
    
    # 计算总计数据
    total_stats = defaultdict(lambda: {
        'messages': 0,
        'chars': 0,
        'ooc_messages': 0,
        'ooc_chars': 0
    })
    for day_stats in daily_stats.values():
        for speaker, stats in day_stats.items():
            total_stats[speaker]['messages'] += stats['messages']
            total_stats[speaker]['chars'] += stats['chars']
            total_stats[speaker]['ooc_messages'] += stats['ooc_messages']
            total_stats[speaker]['ooc_chars'] += stats['ooc_chars']
    
    # 添加总体统计数据
    row = 4
    for speaker, stats in sorted(total_stats.items(), 
                              key=lambda x: x[1]['messages'], 
                              reverse=True):
        avg_chars = stats['chars'] / stats['messages'] if stats['messages'] > 0 else 0
        
        ws_summary.cell(row=row, column=1).value = speaker
        ws_summary.cell(row=row, column=2).value = stats['messages']
        ws_summary.cell(row=row, column=3).value = stats['chars']
        ws_summary.cell(row=row, column=4).value = f"{avg_chars:.1f}"
        ws_summary.cell(row=row, column=5).value = stats['ooc_messages']
        ws_summary.cell(row=row, column=6).value = stats['ooc_chars']
        
        # 添加边框和居中
        for col in range(1, 7):
            cell = ws_summary.cell(row=row, column=col)
            cell.border = border
            if col > 1:  # 第一列左对齐，其他居中
                cell.alignment = centered
        
        row += 1

    # 添加空行
    row += 2
    
    # 在同一工作表中添加每日统计数据
    for day, day_stats in sorted(daily_stats.items()):
        # 添加每日标题
        day_title_cell = ws_summary.cell(row=row, column=1)
        day_title_cell.value = f"第{day}天对话统计"
        day_title_cell.font = title_font
        ws_summary.merge_cells(f'A{row}:F{row}')
        day_title_cell.alignment = Alignment(horizontal='center')
        row += 1
        
        # 添加表头
        day_headers = ["角色名", "发言数", "字数", "平均字数", "场外次数", "场外字数"]
        for col, header in enumerate(day_headers, 1):
            cell = ws_summary.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = day_header_fill
            cell.alignment = centered
            cell.border = border
        row += 1
        
        # 添加数据
        daily_total = defaultdict(int)
        
        for speaker, stats in sorted(day_stats.items(), 
                                  key=lambda x: x[1]['messages'], 
                                  reverse=True):
            avg_chars = stats['chars'] / stats['messages'] if stats['messages'] > 0 else 0
            
            ws_summary.cell(row=row, column=1).value = speaker
            ws_summary.cell(row=row, column=2).value = stats['messages']
            ws_summary.cell(row=row, column=3).value = stats['chars']
            ws_summary.cell(row=row, column=4).value = f"{avg_chars:.1f}"
            ws_summary.cell(row=row, column=5).value = stats['ooc_messages']
            ws_summary.cell(row=row, column=6).value = stats['ooc_chars']
            
            # 添加边框和居中
            for col in range(1, 7):
                cell = ws_summary.cell(row=row, column=col)
                cell.border = border
                if col > 1:  # 第一列左对齐，其他居中
                    cell.alignment = centered
            
            # 累计当日总数
            daily_total['messages'] += stats['messages']
            daily_total['chars'] += stats['chars']
            daily_total['ooc_messages'] += stats['ooc_messages']
            daily_total['ooc_chars'] += stats['ooc_chars']
            
            row += 1
        
        # 添加当日总计行
        avg_chars_total = daily_total['chars'] / daily_total['messages'] if daily_total['messages'] > 0 else 0
        
        ws_summary.cell(row=row, column=1).value = "当日总计"
        ws_summary.cell(row=row, column=1).font = header_font
        ws_summary.cell(row=row, column=2).value = daily_total['messages']
        ws_summary.cell(row=row, column=3).value = daily_total['chars']
        ws_summary.cell(row=row, column=4).value = f"{avg_chars_total:.1f}"
        ws_summary.cell(row=row, column=5).value = daily_total['ooc_messages']
        ws_summary.cell(row=row, column=6).value = daily_total['ooc_chars']
        
        # 添加边框和背景色
        for col in range(1, 7):
            cell = ws_summary.cell(row=row, column=col)
            cell.border = border
            cell.fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
            if col > 1:  # 第一列左对齐，其他居中
                cell.alignment = centered
        
        # 添加空行
        row += 2
    
    # 为每个角色创建对话内容工作表
    for speaker, dialogues in sorted(dialogue_contents.items()):
        print(f"Processing speaker: {speaker}")
        if not dialogues:  # 跳过没有对话的角色
            continue
            
        # 创建工作表（名称不能超过31个字符，且不能包含特殊字符）
        safe_name = re.sub(r'[\[\]\:\*\?\/\\]', '', speaker)[:31]
        ws = wb.create_sheet(title=safe_name)
        
        # 设置列宽
        ws.column_dimensions['A'].width = 40  # 日期/时间戳
        ws.column_dimensions['B'].width = 70  # 对话内容
        
        # 添加标题
        ws['A1'] = f"{speaker} 对话内容"
        ws['A1'].font = title_font
        ws.merge_cells('A1:B1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # 添加表头
        headers = ["时间", "对话内容"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = centered
            cell.border = border


        # 添加对话数据（按天和时间排序）
        row = 4
        for dialogue in sorted(dialogues, key=lambda x: (x['day'], x.get('timestamp', ''))):
            # 时间戳
            timestamp_cell = ws.cell(row=row, column=1)
            timestamp_cell.value = dialogue.get('timestamp', f"第{dialogue['day']}天")
            timestamp_cell.border = border
            timestamp_cell.alignment = centered
            
            # 对话内容
            content_cell = ws.cell(row=row, column=2)
            content_cell.value = dialogue['content']
            content_cell.border = border
            content_cell.alignment = wrapped_text
            
            # 如果是场外内容，使用不同背景色
            if dialogue['is_ooc']:
                timestamp_cell.fill = ooc_fill
                content_cell.fill = ooc_fill
            
            row += 1
    
    # 保存Excel文件
    try:
        wb.save(file_path)
        print(f"Excel文件已生成：{file_path}")
    except Exception as e:
        print(f"生成Excel文件时出错：{e}")

def generate_csv(daily_stats, dialogue_contents, output_file="对话统计.csv"):
    """生成统计数据的CSV文件，并输出Excel文件"""
    total_stats = defaultdict(lambda: {
        'messages': 0,
        'chars': 0,
        'ooc_messages': 0,
        'ooc_chars': 0
    })
    
    # 计算总计数据
    for day_stats in daily_stats.values():
        for speaker, stats in day_stats.items():
            total_stats[speaker]['messages'] += stats['messages']
            total_stats[speaker]['chars'] += stats['chars']
            total_stats[speaker]['ooc_messages'] += stats['ooc_messages']
            total_stats[speaker]['ooc_chars'] += stats['ooc_chars']
    
    try:
        # 生成CSV文件
        with open(output_file, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.writer(csvfile)
            
            # 写入总体统计数据
            writer.writerow(['总体统计'])
            writer.writerow(['角色名', '总发言数', '总字数', '平均字数', '场外次数', '场外字数'])
            
            for speaker, stats in sorted(total_stats.items(), 
                                      key=lambda x: x[1]['messages'], 
                                      reverse=True):
                avg_chars = stats['chars'] / stats['messages'] if stats['messages'] > 0 else 0
                writer.writerow([
                    speaker, 
                    stats['messages'], 
                    stats['chars'], 
                    f"{avg_chars:.1f}", 
                    stats['ooc_messages'], 
                    stats['ooc_chars']
                ])
            
            # 添加空行
            writer.writerow([])
            
            # 写入按日统计数据
            for day, day_stats in sorted(daily_stats.items()):
                writer.writerow([f'第{day}天统计'])
                writer.writerow(['角色名', '发言数', '字数', '平均字数', '场外次数', '场外字数'])
                
                daily_total = defaultdict(int)
                
                for speaker, stats in sorted(day_stats.items(), 
                                         key=lambda x: x[1]['messages'], 
                                         reverse=True):
                    avg_chars = stats['chars'] / stats['messages'] if stats['messages'] > 0 else 0
                    writer.writerow([
                        speaker, 
                        stats['messages'], 
                        stats['chars'], 
                        f"{avg_chars:.1f}", 
                        stats['ooc_messages'], 
                        stats['ooc_chars']
                    ])
                    
                    # 累计当日总数
                    daily_total['messages'] += stats['messages']
                    daily_total['chars'] += stats['chars']
                    daily_total['ooc_messages'] += stats['ooc_messages']
                    daily_total['ooc_chars'] += stats['ooc_chars']
                
                # 写入当日总计
                avg_chars_total = daily_total['chars'] / daily_total['messages'] if daily_total['messages'] > 0 else 0
                writer.writerow([
                    '当日总计', 
                    daily_total['messages'], 
                    daily_total['chars'], 
                    f"{avg_chars_total:.1f}", 
                    daily_total['ooc_messages'], 
                    daily_total['ooc_chars']
                ])
                
                # 添加空行
                writer.writerow([])
        
        print(f"CSV文件已生成：{output_file}")
        
        # 生成Excel文件，包含统计数据和对话内容
        if output_file.endswith('.csv'):
            excel_file = output_file.replace('.csv', '.xlsx')
        else:
            excel_file = output_file + '.xlsx'
        
        generate_excel(daily_stats, dialogue_contents, excel_file)
        
    except Exception as e:
        print(f"生成文件时出错：{e}")

def print_summary_stats(daily_stats):
    """打印汇总统计数据"""
    total_stats = defaultdict(lambda: {
        'messages': 0,
        'chars': 0,
        'ooc_messages': 0,
        'ooc_chars': 0
    })
    
    # 计算总计数据
    for day_stats in daily_stats.values():
        for speaker, stats in day_stats.items():
            total_stats[speaker]['messages'] += stats['messages']
            total_stats[speaker]['chars'] += stats['chars']
            total_stats[speaker]['ooc_messages'] += stats['ooc_messages']
            total_stats[speaker]['ooc_chars'] += stats['ooc_chars']
    
    # 创建简单表格，考虑中文字符宽度
    def get_display_width(s):
        """计算字符串在终端中的显示宽度（汉字占2宽度）"""
        width = 0
        for char in s:
            if '\u4e00' <= char <= '\u9fff':  # 判断是否为汉字
                width += 2
            else:
                width += 1
        return width
    
    def pad_string(s, width, align='left'):
        """填充字符串到指定显示宽度"""
        s_str = str(s)
        display_width = get_display_width(s_str)
        
        # 如果字符串显示宽度超过列宽，截断并添加省略号
        if display_width > width:
            # 逐个字符累加直到接近限制宽度
            truncated = ""
            curr_width = 0
            for char in s_str:
                char_width = 2 if '\u4e00' <= char <= '\u9fff' else 1
                if curr_width + char_width + 3 <= width:  # 为省略号预留宽度
                    truncated += char
                    curr_width += char_width
                else:
                    break
            s_str = truncated + "..."
            display_width = get_display_width(s_str)
        
        if align == 'left':
            return s_str + ' ' * (width - display_width)
        else:  # right align
            return ' ' * (width - display_width) + s_str
    
    # 表头
    name_width = 30  # 角色名列宽增加到30
    num_width = 10   # 数字列宽
    
    print("=" * 80)  # 增加分隔线长度以适应更宽的表格
    print("角色对话统计")
    print("=" * 80)
    print(pad_string('角色名', name_width) + ' '.join(pad_string(col, num_width, 'right') for col in ['总发言数', '总字数', '平均字数', '场外次数', '场外字数']))
    print("-" * 80)
    
    # 按总发言数排序并打印数据
    for speaker, stats in sorted(total_stats.items(), 
                               key=lambda x: x[1]['messages'], 
                               reverse=True):
        avg_chars = stats['chars'] / stats['messages'] if stats['messages'] > 0 else 0
        print(pad_string(speaker, name_width) + 
              pad_string(str(stats['messages']), num_width, 'right') + 
              pad_string(str(stats['chars']), num_width, 'right') + 
              pad_string(f"{avg_chars:.1f}", num_width, 'right') + 
              pad_string(str(stats['ooc_messages']), num_width, 'right') + 
              pad_string(str(stats['ooc_chars']), num_width, 'right'))
    
    print("\n# 按日统计")
    # 输出每日统计
    for day, day_stats in sorted(daily_stats.items()):
        print(f"\n--- 第{day}天 ---")
        print(pad_string('角色名', name_width) + ' '.join(pad_string(col, num_width, 'right') for col in ['发言数', '字数', '平均字数', '场外次数', '场外字数']))
        print("-" * 80)
        
        daily_total = defaultdict(int)
        
        for speaker, stats in sorted(day_stats.items(), 
                                 key=lambda x: x[1]['messages'], 
                                 reverse=True):
            avg_chars = stats['chars'] / stats['messages'] if stats['messages'] > 0 else 0
            print(pad_string(speaker, name_width) + 
                pad_string(str(stats['messages']), num_width, 'right') + 
                pad_string(str(stats['chars']), num_width, 'right') + 
                pad_string(f"{avg_chars:.1f}", num_width, 'right') + 
                pad_string(str(stats['ooc_messages']), num_width, 'right') + 
                pad_string(str(stats['ooc_chars']), num_width, 'right'))
            
            # 累计当日总数
            daily_total['messages'] += stats['messages']
            daily_total['chars'] += stats['chars']
            daily_total['ooc_messages'] += stats['ooc_messages']
            daily_total['ooc_chars'] += stats['ooc_chars']
        
        # 打印当日总计
        print("-" * 80)
        avg_chars_total = daily_total['chars'] / daily_total['messages'] if daily_total['messages'] > 0 else 0
        print(pad_string("当日总计", name_width) + 
              pad_string(str(daily_total['messages']), num_width, 'right') + 
              pad_string(str(daily_total['chars']), num_width, 'right') + 
              pad_string(f"{avg_chars_total:.1f}", num_width, 'right') + 
              pad_string(str(daily_total['ooc_messages']), num_width, 'right') + 
              pad_string(str(daily_total['ooc_chars']), num_width, 'right'))
    
    # CSV格式
    print("\n# CSV格式:")
    print("角色名,总发言数,总字数,平均字数,场外次数,场外字数")
    for speaker, stats in sorted(total_stats.items(), 
                               key=lambda x: x[1]['messages'], 
                               reverse=True):
        avg_chars = stats['chars'] / stats['messages'] if stats['messages'] > 0 else 0
        print(f"{speaker},{stats['messages']},{stats['chars']},{avg_chars:.1f},{stats['ooc_messages']},{stats['ooc_chars']}")

def get_safe_input(prompt, default_value):
    """尝试获取用户输入，失败时使用默认值"""
    try:
        return input(prompt) or default_value
    except Exception as e:
        print(f"无法获取用户输入: {e}，使用默认值: {default_value}")
        return default_value

def find_log_files():
    """在当前目录和logs目录查找可能的日志文件"""
    log_files = []
    
    # 要排除的明显不是日志文件的文件名列表
    exclude_files = ["requirements.txt", "README.md", "setup.py"]
    
    # 检查当前目录
    for file in os.listdir('.'):
        if file.endswith('.txt') and os.path.isfile(file) and file not in exclude_files:
            log_files.append(file)
    
    # 检查logs目录（如果存在）
    logs_dir = "./logs"
    if not os.path.exists(logs_dir):
        try:
            os.makedirs(logs_dir)
            print(f"已创建logs目录: {logs_dir}")
        except Exception as e:
            print(f"创建logs目录失败: {e}")
    
    if os.path.exists(logs_dir) and os.path.isdir(logs_dir):
        for file in os.listdir(logs_dir):
            if file.endswith('.txt') and os.path.isfile(os.path.join(logs_dir, file)):
                log_files.append(os.path.join(logs_dir, file))
    
    return log_files

if __name__ == "__main__":
    print("=" * 70)
    print("欢迎使用角色对话统计工具")
    print("=" * 70)
    print("\n提示: 您可以将日志文件放在当前目录或'logs'文件夹中")
    
    # 检查当前目录下的日志文件
    log_files = find_log_files()
    if log_files:
        print("\n发现可能的日志文件:")
        for i, file in enumerate(log_files, 1):
            print(f"{i}. {file}")
        print("请输入文件编号或直接输入完整文件路径")
    else:
        print("\n未发现日志文件，请将日志文件(.txt)放入当前目录或'logs'文件夹中")
    
    # 执行统计
    try:
        file_input = get_safe_input("请输入日志文件路径（默认为log.txt）：", "log.txt")
        
        # 如果输入的是数字，尝试从发现的文件列表中选择
        if log_files and file_input.isdigit():
            index = int(file_input) - 1
            if 0 <= index < len(log_files):
                file_path = log_files[index]
            else:
                file_path = "log.txt"
        else:
            file_path = file_input
            
        output_file = get_safe_input("请输入输出CSV文件路径（默认为对话统计.csv）：", "对话统计.csv")
        
        print(f"\n正在处理文件：{file_path}")
        daily_stats, dialogue_contents = count_dialogue_from_file(file_path)
        print_summary_stats(daily_stats)
        generate_csv(daily_stats, dialogue_contents, output_file)
        
        print("\n处理完成！按任意键退出...")
        try:
            input()
        except:
            pass
    except Exception as e:
        print(f"程序发生错误: {e}")
        print("按任意键退出...")
        try:
            input()
        except:
            pass